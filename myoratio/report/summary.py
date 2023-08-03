import os
from enum import Enum

import xlsxwriter
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from myoratio.api import Constants
from myoratio.api.helpers import StringHelper, XLSXHelper
from myoratio.stage import Stage


class ParticipantType(Enum):
    HEALTHY = "s"
    WALKER = "m"
    NON_WALKER = "n"


class Summary:
    def __init__(self, data_path: str, analysis: str, stage: str, config: dict) -> None:
        self._data_path = data_path
        self._analysis = analysis

        if stage in [stage.value for stage in Stage]:
            self._stage = stage
        else:
            raise ValueError(f"Expected a value from Stage, but got {stage}")

        self._config = config

        if len(self._config["stages"][self._stage]["label"]) > 0:
            self._stage_folder_name = self._config["stages"][self._stage]["label"]
        else:
            self._stage_folder_name = self._stage

        self._antagonist = self._config["muscles"]["antagonist"]
        self._agonist = self._config["muscles"]["agonist"]

        output_base_path = os.path.join(self._data_path, "Results")

        self._input_base_path = os.path.join(
            output_base_path, self._analysis, self._stage_folder_name
        )

        self._workbook = xlsxwriter.Workbook(
            os.path.join(
                output_base_path,
                f"summary_{self._analysis}_{self._stage_folder_name}.xlsx",
            )
        )

        self._summary_worksheet = self._workbook.add_worksheet("summary")
        self._formats = XLSXHelper.get_cell_formats(self._workbook)

    def _get_headers_by_participant_type(self, participant_type: str) -> list[str]:
        headers = ["Amplitude (°)", "Duration (s)", "Ratio"]

        if participant_type == ParticipantType.HEALTHY.value:
            headers.extend(["Healthy Participant"])
        elif participant_type == ParticipantType.WALKER.value:
            headers.extend(["Walking Participant"])
        elif participant_type == ParticipantType.NON_WALKER.value:
            headers.extend(["Non-walking Participant"])
        else:
            raise ValueError(
                f"Expected a value from ParticipantType, but got {participant_type}"
            )

        return list(reversed(headers))

    def _get_total_number_iterations(self, sheet: Worksheet) -> int:
        row_index = 0

        for row in sheet.iter_rows():
            cell_value = str(row[0].value).lower()

            if cell_value == "average":
                row_index = row[0].row
                break

        return row_index - Constants.XSLX_TABLE_SPACES.value

    def _get_ratio_value(self, sheet: Worksheet) -> float:
        target_row_index = 0

        for row in sheet.iter_rows():
            cell_value = str(row[0].value).lower()

            if "average ratios" in cell_value:
                target_row_index = row[0].row
                break

        column_ratio_index = None
        row_ratio_index = None

        for column in sheet.iter_cols(min_row=target_row_index, max_row=target_row_index):
            cell_value = str(column[0].value).lower()

            if self._agonist in cell_value:
                column_ratio_index = column[0].column
                break

        for row in sheet.iter_rows(
            min_row=target_row_index + 1,
            min_col=Constants.XSLX_TABLE_SPACES.value,
            max_col=Constants.XSLX_TABLE_SPACES.value,
        ):
            cell_value = str(row[0].value).lower()

            if self._antagonist in cell_value:
                row_ratio_index = row[0].row
                break

        if row_ratio_index is None and column_ratio_index is None:
            raise ValueError("Ratio cell cannot be found")

        return float(
            sheet.cell(
                row=row_ratio_index, column=column_ratio_index  # type: ignore
            ).value
        )

    def _get_duration_or_amplitude_value(
        self,
        sheet: Worksheet,
        total_iterations: int,
        duration: bool = True,
        amplitude: bool = False,
    ) -> float:
        column_start_index = None
        column_stop_index = None

        for column in sheet.iter_cols():
            cell_value = str(column[0].value).lower()

            start_substring = (
                "start time" if duration and not amplitude else "start angle"
            )
            stop_substring = "stop time" if duration and not amplitude else "stop angle"

            if start_substring in cell_value:
                column_start_index = column[0].column
            elif stop_substring in cell_value:
                column_stop_index = column[0].column

            if column_start_index is not None and column_stop_index is not None:
                break

        if column_start_index is None and column_stop_index is None:
            raise ValueError("Start and stop angles columns cannot be found")

        value = 0

        for i in range(total_iterations):
            if duration and not amplitude:
                value += float(
                    str(
                        sheet.cell(
                            row=i + Constants.XSLX_TABLE_SPACES.value,
                            column=column_stop_index,  # type: ignore
                        ).value
                    )
                ) - float(
                    str(
                        sheet.cell(
                            row=i + Constants.XSLX_TABLE_SPACES.value,
                            column=column_start_index,  # type: ignore
                        ).value
                    )
                )

            else:
                value += abs(
                    float(
                        str(
                            sheet.cell(
                                row=i + Constants.XSLX_TABLE_SPACES.value,
                                column=column_start_index,  # type: ignore
                            ).value
                        )
                    )
                    - float(
                        str(
                            sheet.cell(
                                row=i + Constants.XSLX_TABLE_SPACES.value,
                                column=column_stop_index,  # type: ignore
                            ).value
                        )
                    )
                )

        return value / total_iterations

    def _get_duration_value(self, sheet: Worksheet, total_iterations: int) -> float:
        return self._get_duration_or_amplitude_value(
            sheet, total_iterations, duration=True, amplitude=False
        )

    def _get_amplitude_value(self, sheet: Worksheet, total_iterations: int) -> float:
        return self._get_duration_or_amplitude_value(
            sheet, total_iterations, duration=False, amplitude=True
        )

    def generate_XLSX_summary(self) -> None:
        input_path = os.path.join(
            self._data_path,
            "Results",
            self._analysis,
            self._stage_folder_name,
        )

        report_files = os.listdir(input_path)

        if len(report_files) <= 0:
            raise FileNotFoundError(f"Path {input_path} does not contain any reports")

        sorted_report_files = sorted(
            os.listdir(input_path), key=lambda x: int(x.split("_")[-2])
        )

        participtant_type_processed = []

        current_row = 0

        for report_file in sorted_report_files:
            participant_type = os.path.splitext(report_file)[0][-1].lower()

            if participant_type not in participtant_type_processed:
                if len(participtant_type_processed) >= 1:
                    current_row += Constants.XSLX_TABLE_SPACES.value
                else:
                    current_row += 1

                participtant_type_processed.extend([participant_type])
                headers = self._get_headers_by_participant_type(participant_type)

                self._summary_worksheet.write_row(
                    f"A{current_row}", headers, self._formats["header"]
                )

                current_row += 1

            report_file_path = os.path.join(self._input_base_path, report_file)
            report_workbook = load_workbook(report_file_path)
            report_sheet = report_workbook["report"]

            participant_data = []
            participant = StringHelper.extract_participant_from_report_filename(
                report_file
            )

            participant_data.extend([participant])

            total_iterations = self._get_total_number_iterations(report_sheet)

            ratio_value = self._get_ratio_value(report_sheet)
            participant_data.extend([ratio_value])

            duration_value = self._get_duration_value(report_sheet, total_iterations)
            participant_data.extend([duration_value])

            amplitude_value = self._get_amplitude_value(report_sheet, total_iterations)
            participant_data.extend([amplitude_value])

            self._summary_worksheet.write_row(f"A{current_row}", participant_data, None)
            self._summary_worksheet.set_column(0, 0, None, self._formats["center"])
            self._summary_worksheet.set_column(1, 1, None, self._formats["number_short"])
            self._summary_worksheet.set_column(2, 3, None, self._formats["number"])

            current_row += 1

        self._summary_worksheet.autofit()
        self._workbook.close()
